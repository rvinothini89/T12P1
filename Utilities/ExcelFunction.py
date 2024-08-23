# using openpyxl module for working with excel sheets
from openpyxl import load_workbook
# using datetime module for getting current time for updating execution time in excel
from datetime import datetime


class excelFunction:
    username = None
    password = None

    # initializing class with excel file and its respective sheet number where test cases are present
    def __init__(self, file_name, sheet_number):
        self.file = file_name
        self.sheet = sheet_number

    # fetch the row count of the Excel file
    def row_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_row

    # fetch the column count of the Excel file
    def column_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_column

    # read the data from the Excel file
    def read_data(self, row_number, column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.cell(row=row_number, column=column_number).value

    # write the data to the Excel file
    def write_data(self, row_number, column_number, data, error=None):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet.cell(row=row_number, column=column_number).value = data
        # in case of error, this if will take care of printing the error in excel
        if error:
            sheet.cell(row=row_number, column=column_number + 1).value = error
        # this will take care of updating execution time in the excel sheet
        execution_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.cell(row=row_number, column=4).value = execution_time
        workbook.save(self.file)